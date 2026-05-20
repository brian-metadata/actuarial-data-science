"""
Generate Metadata Master Sheet v2 2026 Excel file
Based on specification in readme-plan.md
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# File name
output_file = 'Metadata_Master_Sheet_v2_2026.xlsx'

# ============================================================================
# TAB 1: Sources (SOR Inventory – Raw Focus)
# ============================================================================
sources_columns = [
    'SOR_Name',
    'Source_Table_or_View_Name',
    'Database_or_Schema',
    'Approx_Daily_Records_or_Size',
    'Ingestion_Pattern',
    'Key_Columns_with_Types',
    'Primary_Key_or_Unique_ID',
    'Business_Meanings_Key_Terms',
    'PII_or_Sensitive_Flag',
    'Current_Use_Reports_or_Rate_Cases',
    'SME_Notes_Questions_Gaps',
    'Filled_By_Date'
]

sources_data = [
    {
        'SOR_Name': 'Duck Creek Clarity',
        'Source_Table_or_View_Name': 'POLICY_CURATED',
        'Database_or_Schema': 'CLARITY_DB.POLICY_LAYER',
        'Approx_Daily_Records_or_Size': '~50k recs / 200 MB daily',
        'Ingestion_Pattern': 'Batch Export or API',
        'Key_Columns_with_Types': 'PolicyGUID:VARCHAR, EffectiveDate:DATE, Premium:DECIMAL(12,2), PolicyNumber:VARCHAR',
        'Primary_Key_or_Unique_ID': 'PolicyGUID',
        'Business_Meanings_Key_Terms': 'PolicyGUID = Unique system ID; Premium = Annual amount; EffectiveDate = Policy start date',
        'PII_or_Sensitive_Flag': 'Yes – InsuredName, Address',
        'Current_Use_Reports_or_Rate_Cases': 'Rate case exposure by state; Sales Summary',
        'SME_Notes_Questions_Gaps': 'History in separate view? Deletes?',
        'Filled_By_Date': 'Jane Doe – 2026-02-05'
    },
    {
        'SOR_Name': 'Guidewire',
        'Source_Table_or_View_Name': 'POLICY_CDC_PARQUET',
        'Database_or_Schema': 'GW_CDC_EXPORT',
        'Approx_Daily_Records_or_Size': '~100k recs / 500 MB daily',
        'Ingestion_Pattern': 'CDC',
        'Key_Columns_with_Types': 'PolicyID:VARCHAR, TransactionDate:TIMESTAMP, ChangeType:VARCHAR, PolicyData:STRUCT',
        'Primary_Key_or_Unique_ID': 'PolicyID + TransactionDate',
        'Business_Meanings_Key_Terms': 'PolicyID = Guidewire policy identifier; ChangeType = INSERT/UPDATE/DELETE; PolicyData = Full policy JSON',
        'PII_or_Sensitive_Flag': 'Yes – Customer data in PolicyData struct',
        'Current_Use_Reports_or_Rate_Cases': 'Real-time policy change tracking; Audit reports',
        'SME_Notes_Questions_Gaps': 'Auto Loader compatible?',
        'Filled_By_Date': 'John Smith – 2026-02-06'
    }
]

sources_df = pd.DataFrame(sources_data, columns=sources_columns)

# ============================================================================
# TAB 2: Targets (Bronze / Silver Placeholders)
# ============================================================================
targets_columns = [
    'SOR_Name',
    'Source_Table_or_View',
    'Bronze_Table_Name',
    'Bronze_Schema_Simple',
    'Append_Only',
    'Silver_Table_Name',
    'Unity_Catalog_Path',
    'Gap_Flag',
    'Notes'
]

targets_data = [
    {
        'SOR_Name': 'Duck Creek Clarity',
        'Source_Table_or_View': 'POLICY_CURATED',
        'Bronze_Table_Name': 'bronze_duck_creek_policy_curated',
        'Bronze_Schema_Simple': 'PolicyGUID:VARCHAR, EffectiveDate:DATE, Premium:DECIMAL(12,2)',
        'Append_Only': 'Yes',
        'Silver_Table_Name': 'silver_duck_creek_policy_cleaned',
        'Unity_Catalog_Path': '/catalog/bronze/duck_creek_policy',
        'Gap_Flag': 'No',
        'Notes': ''
    },
    {
        'SOR_Name': 'Guidewire',
        'Source_Table_or_View': 'POLICY_CDC_PARQUET',
        'Bronze_Table_Name': 'bronze_guidewire_policy_cdc',
        'Bronze_Schema_Simple': 'PolicyID:VARCHAR, TransactionDate:TIMESTAMP, ChangeType:VARCHAR',
        'Append_Only': 'Yes',
        'Silver_Table_Name': 'silver_guidewire_policy_enriched',
        'Unity_Catalog_Path': '/catalog/bronze/guidewire_policy',
        'Gap_Flag': 'No',
        'Notes': 'Auto Loader pattern'
    }
]

targets_df = pd.DataFrame(targets_data, columns=targets_columns)

# ============================================================================
# TAB 3: Pipelines (High-Level Definitions)
# ============================================================================
pipelines_columns = [
    'Pipeline_ID',
    'SOR_Name',
    'Source_Table',
    'Bronze_Table',
    'Pattern',
    'iPaaS_Required',
    'Estimated_Effort',
    'Risks_AntiPattern',
    'Status'
]

pipelines_data = [
    {
        'Pipeline_ID': 'DC_CLARITY_POLICY_BATCH',
        'SOR_Name': 'Duck Creek Clarity',
        'Source_Table': 'POLICY_CURATED',
        'Bronze_Table': 'bronze_duck_creek_policy_curated',
        'Pattern': 'Batch Export',
        'iPaaS_Required': 'Yes (API mediation)',
        'Estimated_Effort': 'Medium',
        'Risks_AntiPattern': 'Partner constraints',
        'Status': 'Planned'
    },
    {
        'Pipeline_ID': 'GW_POLICY_CDC',
        'SOR_Name': 'Guidewire',
        'Source_Table': 'POLICY_CDC_PARQUET',
        'Bronze_Table': 'bronze_guidewire_policy_cdc',
        'Pattern': 'CDC',
        'iPaaS_Required': 'No',
        'Estimated_Effort': 'Low',
        'Risks_AntiPattern': '',
        'Status': 'Planned'
    }
]

pipelines_df = pd.DataFrame(pipelines_data, columns=pipelines_columns)

# ============================================================================
# TAB 4: Gaps & Actions
# ============================================================================
gaps_columns = [
    'Area',
    'Gap Description',
    'Type (Strategic/Tactical)',
    'Priority',
    'Status',
    'Action Plan',
    'Owner',
    'Deadline',
    'Notes'
]

gaps_data = [
    {
        'Area': 'Metadata Collection',
        'Gap Description': 'DDs not pulled from vendor portals',
        'Type (Strategic/Tactical)': 'Tactical',
        'Priority': 'High',
        'Status': 'Open – needs closure',
        'Action Plan': 'Delegate DD pulls: Duck Creek Support Portal, Guidewire Data Studio',
        'Owner': 'SOR SMEs',
        'Deadline': '2026-02-14',
        'Notes': 'Blocking pipeline design'
    },
    {
        'Area': 'Glossary',
        'Gap Description': 'No centralized BA/architect glossary; mapping needed',
        'Type (Strategic/Tactical)': 'Strategic',
        'Priority': 'Medium',
        'Status': 'In Progress',
        'Action Plan': 'Seed from MVM Business Meanings column',
        'Owner': 'Cognizant / BA',
        'Deadline': '2026-02-21',
        'Notes': 'Foundation for KG'
    },
    {
        'Area': 'iPaaS Flow Mandates',
        'Gap Description': 'BU relation to iPaaS undefined; why #3?',
        'Type (Strategic/Tactical)': 'Tactical',
        'Priority': 'High',
        'Status': 'Open – needs closure',
        'Action Plan': 'Define BU/iPaaS mandates via questionnaire',
        'Owner': 'iPaaS SME / BU',
        'Deadline': '2026-02-14',
        'Notes': 'Core triples gap'
    },
    {
        'Area': 'Triples Problem Flows',
        'Gap Description': 'Why OLTP feedback if SoR → lakehouse covered?',
        'Type (Strategic/Tactical)': 'Tactical',
        'Priority': 'High',
        'Status': 'Open – core gap',
        'Action Plan': 'Map #1–#3 per pipeline; justify #3 loops',
        'Owner': 'iPaaS / BU SMEs',
        'Deadline': '2026-02-14',
        'Notes': 'Analytical vs operational needs'
    }
]

gaps_df = pd.DataFrame(gaps_data, columns=gaps_columns)

# ============================================================================
# TAB 5: Glossary (Seeded from MVM)
# ============================================================================
glossary_columns = [
    'Term',
    'Definition',
    'Category',
    'Used_In',
    'Source'
]

glossary_data = [
    {
        'Term': 'PolicyGUID',
        'Definition': 'Unique system-generated policy identifier',
        'Category': 'Field Meaning',
        'Used_In': 'Duck Creek Clarity – POLICY_CURATED',
        'Source': 'Sources Tab – Row 1'
    },
    {
        'Term': 'Premium',
        'Definition': 'Annual policy premium amount',
        'Category': 'Field Meaning',
        'Used_In': 'Duck Creek Clarity – POLICY_CURATED',
        'Source': 'Sources Tab – Row 1'
    },
    {
        'Term': 'CDC',
        'Definition': 'Change Data Capture - incremental data extraction pattern',
        'Category': 'Ingestion Pattern',
        'Used_In': 'Guidewire – POLICY_CDC_PARQUET',
        'Source': 'Sources Tab – Row 2'
    },
    {
        'Term': 'Bronze Layer',
        'Definition': 'Raw append-only data layer in medallion architecture',
        'Category': 'Architecture',
        'Used_In': 'All SORs',
        'Source': 'Targets Tab'
    }
]

glossary_df = pd.DataFrame(glossary_data, columns=glossary_columns)

# ============================================================================
# TAB 6: Gap Analysis (Current vs Future State)
# ============================================================================
gap_analysis_columns = [
    'Area',
    'Current State',
    'Future State',
    'Gap Description',
    'Type (Strategic/Tactical)',
    'Priority',
    'Action Plan',
    'Owner',
    'Deadline'
]

gap_analysis_data = [
    {
        'Area': 'Metadata Collection',
        'Current State': 'Messy Cognizant Excels, partial DDs, no structure',
        'Future State': 'Structured MVM sheets per SOR, feeding KG',
        'Gap Description': 'DDs not pulled from vendor portals',
        'Type (Strategic/Tactical)': 'Tactical',
        'Priority': 'High',
        'Action Plan': 'Delegate DD pulls: Duck Creek Support Portal, Guidewire Data Studio',
        'Owner': 'SOR SMEs',
        'Deadline': '2026-02-14'
    },
    {
        'Area': 'Glossary',
        'Current State': 'Scattered or none; seeded minimally from MVM',
        'Future State': 'AI-ready KG with full business terms & lineage',
        'Gap Description': 'No centralized BA/architect glossary; mapping needed',
        'Type (Strategic/Tactical)': 'Strategic',
        'Priority': 'Medium',
        'Action Plan': 'Seed from MVM Business Meanings column',
        'Owner': 'Cognizant / BA',
        'Deadline': '2026-02-21'
    },
    {
        'Area': 'iPaaS Flow Mandates',
        'Current State': 'Unclear BU rules; "triples" not mapped',
        'Future State': 'Defined per-SOR authoritative flows (#1–#3)',
        'Gap Description': 'BU relation to iPaaS undefined; why #3?',
        'Type (Strategic/Tactical)': 'Tactical',
        'Priority': 'High',
        'Action Plan': 'Define BU/iPaaS mandates via questionnaire',
        'Owner': 'iPaaS SME / BU',
        'Deadline': '2026-02-14'
    },
    {
        'Area': 'Triples Problem Flows',
        'Current State': 'Unknown authoritative patterns',
        'Future State': 'Mapped #1–#3 per pipeline; justified #3 loops',
        'Gap Description': 'Why OLTP feedback if SoR → lakehouse covered?',
        'Type (Strategic/Tactical)': 'Tactical',
        'Priority': 'High',
        'Action Plan': 'Map #1–#3 per pipeline; justify #3 loops',
        'Owner': 'iPaaS / BU SMEs',
        'Deadline': '2026-02-14'
    }
]

gap_analysis_df = pd.DataFrame(gap_analysis_data, columns=gap_analysis_columns)

# ============================================================================
# TAB 7: Triples Analysis (The Core Gap)
# ============================================================================
triples_columns = [
    'SOR_Pipeline',
    'Flow_1_Direct_to_SoR_Lakehouse',
    'Flow_2_Via_iPaaS_to_SoR',
    'Flow_3_iPaaS_to_OLTP',
    'Why_Flow_3_Needed_If_1_and_2_Cover_SoR',
    'Gap_Status',
    'Type'
]

triples_data = [
    {
        'SOR_Pipeline': 'Duck Creek Clarity API',
        'Flow_1_Direct_to_SoR_Lakehouse': 'No',
        'Flow_2_Via_iPaaS_to_SoR': 'Yes (API mediation, XML retrieval)',
        'Flow_3_iPaaS_to_OLTP': 'Yes (operational triggers for policy updates)',
        'Why_Flow_3_Needed_If_1_and_2_Cover_SoR': 'Lakehouse is OLAP only – OLTP needs real-time sync for app actions',
        'Gap_Status': 'Open',
        'Type': 'Tactical'
    },
    {
        'SOR_Pipeline': 'Guidewire Parquet CDC',
        'Flow_1_Direct_to_SoR_Lakehouse': 'Yes (Direct CDC via Auto Loader)',
        'Flow_2_Via_iPaaS_to_SoR': 'No',
        'Flow_3_iPaaS_to_OLTP': 'No',
        'Why_Flow_3_Needed_If_1_and_2_Cover_SoR': 'N/A - Direct pattern, no iPaaS needed',
        'Gap_Status': 'Closed',
        'Type': 'Tactical'
    }
]

triples_df = pd.DataFrame(triples_data, columns=triples_columns)

# ============================================================================
# CREATE EXCEL FILE
# ============================================================================
print(f"Creating {output_file}...")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    sources_df.to_excel(writer, sheet_name='Sources', index=False)
    targets_df.to_excel(writer, sheet_name='Targets', index=False)
    pipelines_df.to_excel(writer, sheet_name='Pipelines', index=False)
    gaps_df.to_excel(writer, sheet_name='Gaps & Actions', index=False)
    glossary_df.to_excel(writer, sheet_name='Glossary', index=False)
    gap_analysis_df.to_excel(writer, sheet_name='Gap Analysis', index=False)
    triples_df.to_excel(writer, sheet_name='Triples Analysis', index=False)

# ============================================================================
# APPLY FORMATTING
# ============================================================================
print("Applying formatting...")

wb = load_workbook(output_file)

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
wrap_align = Alignment(wrap_text=True, vertical='top')

# Color fills for status/priority
status_colors = {
    'Open': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'Open – needs closure': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'Open – core gap': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    'In Progress': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    'Closed': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'High': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'Medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    'Low': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'Tactical': PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
    'Strategic': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
}

# Format each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply borders and alignment to data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = wrap_align
    
    # Apply conditional formatting for Status/Priority columns
    if sheet_name in ['Gaps & Actions', 'Gap Analysis']:
        # Find Status column
        status_col = None
        priority_col = None
        type_col = None
        
        for idx, cell in enumerate(ws[1], 1):
            if 'Status' in str(cell.value):
                status_col = idx
            if 'Priority' in str(cell.value):
                priority_col = idx
            if 'Type' in str(cell.value) and 'Strategic' in str(cell.value):
                type_col = idx
        
        # Apply fills based on values
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if status_col:
                cell = row[status_col - 1]
                status_val = str(cell.value) if cell.value else ''
                if status_val in status_colors:
                    cell.fill = status_colors[status_val]
            
            if priority_col:
                cell = row[priority_col - 1]
                priority_val = str(cell.value) if cell.value else ''
                if priority_val in status_colors:
                    cell.fill = status_colors[priority_val]
            
            if type_col:
                cell = row[type_col - 1]
                type_val = str(cell.value) if cell.value else ''
                if type_val in status_colors:
                    cell.fill = status_colors[type_val]
    
    # Freeze header row
    ws.freeze_panes = 'A2'

# Save the formatted workbook
wb.save(output_file)
print(f"Successfully created {output_file}")
print(f"  - 7 tabs created with example data")
print(f"  - Headers formatted with blue background")
print(f"  - Conditional formatting applied for Status/Priority")
print(f"  - Column widths auto-adjusted")
print(f"  - Header rows frozen for scrolling")

